#!/usr/bin/python
import sys, os
sys.path.append("../../lib")
from livdatcsvlib import *
import openpyxl
import wave
import contextlib

def getaudiolen(fname):
	with contextlib.closing(wave.open(fname,'r')) as f:
		frames = f.getnframes()
		rate = f.getframerate()
		duration = frames / float(rate)
		return duration


def lookupnamefornumber(number,people):
	for person in people.matrix:
			if person['Mob1'].strip(".0")==number or person['Mob2'].strip(".0")==number:
				return person['Name']
				
if __name__=="__main__":
	callspath=sys.argv[1]
	outfilename=sys.argv[2]
	calls=os.listdir(callspath)
	callkeys=["number","server","year","month","day","hour","min","sec"]
	calllist=[]
	e=ExcelFile()
	e.importascsv("/home/pi/MasterSKSList.xlsx")
	people=e.worksheets[0]
	
	for call in calls:
		if "calllog" in call:
			break
		i=0
		dictionary={}
		callvals=call.split("-")
		for key in callkeys:
			dictionary[callkeys[i]]=callvals[i]
			i+=1
		dictionary['name']='Unknown'
		dictionary['number']=dictionary['number'].replace("+91","")
		dictionary['name']=lookupnamefornumber(dictionary['number'],people)
		dictionary['date']=dictionary['year']+"-"+dictionary['month']+"-"+dictionary['day']
		dictionary['time']=dictionary['hour']+":"+dictionary['min']+":"+dictionary['sec']
		callfiles=os.listdir(os.path.join(callspath,call))
		dictionary['recording']="None"
		dictionary['recordlen']="None"
		for callfile in callfiles:
			if "callfile" in callfile:
				dictionary['recording']=callfile
				dictionary['recordlen']=getaudiolen(os.path.join(callspath,call,callfile))
		calllist.append(dictionary)
	c=CSVFile()
	c.colnames=callkeys
	c.colnames+=["name","date","time","recording","recordlen"]
	c.matrix=calllist
	
	
	
	x=ExcelFile()
	x.worksheetnames.append('March-2017-Calls')
	
	x.worksheets.append(c)
	callsbynum=CSVFile()
	callsbynum.filename="Calls by Number"
	callsbynum.colnames=["Number","Number of Calls","Name of Caller","Number of Recordings"]
	numlist=[]
	for call in c.matrix:
		numlist.append(call['number'].replace("+91",""))
	numlist=list(set(numlist))
	callmatrix=[]
	for num in numlist:
		dictionary={}
		numcalls=0
		numrecs=0
		for call in c.matrix:
			if call['number']==num:
				numcalls+=1
				if call['recording']!="None":
					numrecs+=1
		dictionary['Number']=num
		dictionary['Name of Caller']=lookupnamefornumber(dictionary['Number'],people)
		dictionary['Number of Calls']=numcalls
		dictionary['Number of Recordings']=numrecs
		callmatrix.append(dictionary)
	callsbynum.matrix=callmatrix
	
	x.worksheetnames.append("Calls by Number")
	x.worksheets.append(callsbynum)
	x.exportfile(outfilename)
	
	xfile=openpyxl.load_workbook(outfilename)
	callsbyday=xfile.create_sheet("Calls by Day")
	callsbyday["A1"]="Day"
	callsbyday["B1"]="Number of Calls"
	callsbyday["C1"]="Number of Recordings"
	
	days=[]
	for call in c.matrix:
		days.append(call['date'])
	days=list(set(days))
	days.sort()
	daycountdict=[]
	for day in days:
		daycount=0
		dayrecount=0
		for call in c.matrix:
			if call['date']==day:
				daycount+=1
				if call['recording']!="None":
					dayrecount+=1
		daydict={}
		daydict['Day']=datetime.datetime.strptime(day,"%Y-%b-%d")
		daydict['Number of Calls']=daycount
		daydict["Number of Recordings"]=dayrecount
		
		daycountdict.append(daydict)
	r=2
	c=1
	for daydict in daycountdict:
		cell=callsbyday.cell(row=r,column=c)
		cell.value=daydict['Day']
		c+=1
		cell=callsbyday.cell(row=r,column=c)
		cell.value=daydict['Number of Calls']
		c+=1
		cell=callsbyday.cell(row=r,column=c)
		cell.value=daydict['Number of Recordings']
		
		r+=1
		c=1
	
	
	
	xfile.save(outfilename)
	
		
