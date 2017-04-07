#!/usr/bin/python
import sys, os
sys.path.append("../lib")
from livdatcsvlib import *
import openpyxl

def lookupnamefornumber(number,people):
	for person in people.matrix:
			if person['Mob1'].strip(".0")==number or person['Mob2'].strip(".0")==number:
				return person['Name']
				
if __name__=="__main__":
	callspath=sys.argv[1]
	outfilename=sys.argv[2]
	calls=os.listdir(callspath)
	callkeys=["number","server","year","month","day","hour","min","sec"]
	calllist=[]
	e=ExcelFile()
	e.importascsv("/home/arjun/MasterSKSList.xlsx")
	people=e.worksheets[0]
	
	for call in calls:
		i=0
		dictionary={}
		callvals=call.split("-")
		for key in callkeys:
			dictionary[callkeys[i]]=callvals[i]
			i+=1
		dictionary['name']='Unknown'
		dictionary['number']=dictionary['number'].replace("+91","")
		dictionary['name']=lookupnamefornumber(dictionary['number'],people)
		dictionary['date']=dictionary['year']+"-"+dictionary['month']+"-"+dictionary['day']
		dictionary['time']=dictionary['hour']+":"+dictionary['min']+":"+dictionary['sec']
		calllist.append(dictionary)
	c=CSVFile()
	c.colnames=callkeys
	c.colnames+=["name","date","time"]
	c.matrix=calllist
	
	
	
	x=ExcelFile()
	x.worksheetnames.append('March-2017-Calls')
	
	x.worksheets.append(c)
	callsbynum=CSVFile()
	callsbynum.filename="Calls by Number"
	callsbynum.colnames=["Number","Number of Calls","Name of Caller"]
	numlist=[]
	for call in c.matrix:
		numlist.append(call['number'].replace("+91",""))
	numlist=list(set(numlist))
	callmatrix=[]
	for num in numlist:
		dictionary={}
		numcalls=0
		for call in c.matrix:
			if call['number']==num:
				numcalls+=1
		dictionary['Number']=num
		dictionary['Name of Caller']=lookupnamefornumber(dictionary['Number'],people)
		dictionary['Number of Calls']=numcalls
		callmatrix.append(dictionary)
	callsbynum.matrix=callmatrix
	
	x.worksheetnames.append("Calls by Number")
	x.worksheets.append(callsbynum)
	x.exportfile(outfilename)
	
	xfile=openpyxl.load_workbook(outfilename)
	callsbyday=xfile.create_sheet("Calls by Day")
	callsbyday["A1"]="Day"
	callsbyday["B1"]="Number of Calls"
	days=[]
	for call in c.matrix:
		days.append(call['date'])
	days=list(set(days))
	days.sort()
	daycountdict={}
	for day in days:
		daycount=0
		for call in c.matrix:
			if call['date']==day:
				daycount+=1
		daycountdict[day]=daycount
	r=2
	c=1
	for key,val in daycountdict.iteritems():
		cell=callsbyday.cell(row=r,column=c)
		cell.value=datetime.datetime.strptime(key,"%Y-%b-%d")
		c+=1
		cell=callsbyday.cell(row=r,column=c)
		cell.value=val
		
		r+=1
		c=1
	
	
	
	xfile.save(outfilename)
	
		
